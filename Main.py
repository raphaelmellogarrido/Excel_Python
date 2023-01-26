import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference

workbook = Workbook()

olimpiadas = workbook.create_sheet("olimpíadas", 0)
print(workbook.sheetnames)
print(olimpiadas)
olimpiadas = workbook.active

lista_olimpiadas = [['USA', 26, 12, 5],['China', 38, 20, 7], ['UK', 29, 7, 7], ['Russia', 22, 10, 9], ['South Korea', 13, 3, 2], ['Germany', 11, 7, 4]]

data = (
    (lista_olimpiadas[0][0], lista_olimpiadas[0][1], lista_olimpiadas[0][2], lista_olimpiadas[0][3]),
    (lista_olimpiadas[1][0], lista_olimpiadas[1][1], lista_olimpiadas[1][2], lista_olimpiadas[1][3]),
    (lista_olimpiadas[2][0], lista_olimpiadas[2][1], lista_olimpiadas[2][2], lista_olimpiadas[2][3]),
    (lista_olimpiadas[3][0], lista_olimpiadas[3][1], lista_olimpiadas[3][2], lista_olimpiadas[3][3]),
    (lista_olimpiadas[4][0], lista_olimpiadas[4][1], lista_olimpiadas[4][2], lista_olimpiadas[4][3]),
    (lista_olimpiadas[5][0], lista_olimpiadas[5][1], lista_olimpiadas[5][2], lista_olimpiadas[5][3])
)
for i in data:
    olimpiadas.append(i)

for fila in olimpiadas.rows:
    for cell in fila:
        print(cell.value)
for coluna in olimpiadas.columns:
    for cell in coluna:
        print(cell.value)

olimpiadas.move_range('A1:E1', rows=1)

olimpiadas['A1'] = 'País'
olimpiadas['B1'] = 'Ouros'
olimpiadas['C1'] = 'Pratas'
olimpiadas['D1'] = 'Bronzes'

negrito = Font(bold=True)
for cell in olimpiadas["1:1"]:
    cell.font = negrito

olimpiadas['E1'] = 'Total'
olimpiadas['E2'] = '= SUM(B2:C2:D2)'
olimpiadas['E3'] = '= SUM(B3:C3:D3)'
olimpiadas['E4'] = '= SUM(B4:C4:D4)'
olimpiadas['E5'] = '= SUM(B5:C5:D5)'
olimpiadas['E6'] = '= SUM(B6:C6:D6)'

values = Reference(olimpiadas, min_col=5, max_col=5, min_row=1, max_row=7)
cats = Reference(olimpiadas, min_col=1, min_row=2, max_row=7)
chart = BarChart()
chart.add_data(values, titles_from_data=True)
chart.set_categories(cats)
olimpiadas.add_chart(chart,"F2")

workbook.save('testing.xlsx')
workbook.close()